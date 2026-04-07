import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Questions"

headers = [
    'SubjectName','TopicName','Subtopic','DifficultyLevel','ComplexityLevel',
    'ExamType','Marks','NegativeMarks','QuestionType','QuestionText',
    'OptionA','OptionB','OptionC','OptionD','CorrectOptions',
    'NumericalAnswer','Tags','Solution'
]

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

questions = [
    # ── PASSAGE 1: Miss Stubbs (Q530-533) ────────────────────────────────
    ['English','Reading Comprehension','Passage Comprehension','Medium','Medium','CUET','5','-1','MCQ',
     "Passage: The next stage of the visit began as Mrs. Broadwidth brought in a cup of tea. It was the usual scenario for the many cups of tea I had drunk with Miss Stubbs under the little card which dangled above her bed. She mostly liked to talk about her pets and the ones she had known right back to her girlhood. She spoke a lot about the days her family was alive. She loved to describe the escapades of her three brothers. One day she showed me a photograph which Mrs. Broadwidth had found. The things I had heard in the village came back to me: about the prosperous father and his family who lived in the big house once. When the foreign investments crashed and the sudden change in circumstances, the old father died. Probably just enough brass to keep Miss Stubbs and her animals alive. And, sitting there, I felt the devotion of this shaggy bunch whose eyes were never far from her face.\n\nQ: Miss Stubbs led a very simple life because ___",
     "she didn't believe in leading a luxurious life",
     "she was too sad a person to enjoy simple pleasures of life",
     "she couldn't afford even the normal little luxuries",
     "she was bed-ridden and bitter about it",
     'C','','Comprehension,CUET English 2025,Miss Stubbs','When the investments crashed and the old father died, there was barely enough money to keep going — she could not afford luxuries.'],

    ['English','Reading Comprehension','Passage Comprehension','Medium','Medium','CUET','5','-1','MCQ',
     "Which among the following was NOT a topic of Miss Stubbs's conversation?",
     "her brothers","her pets","her part with her family","her classmates",
     'D','','Comprehension,CUET English 2025,Miss Stubbs','The passage mentions brothers, pets, and family days — not classmates.'],

    ['English','Reading Comprehension','Passage Comprehension','Medium','Medium','CUET','5','-1','MCQ',
     "There's not much \"brass\" here now — the underlined word refers to:",
     "brass ware","musical instruments","money","metal",
     'C','','Comprehension,CUET English 2025,Miss Stubbs','"Brass" is a colloquial/informal word for money.'],

    ['English','Reading Comprehension','Passage Comprehension','Medium','Medium','CUET','5','-1','MCQ',
     "Identify the meaning of the underlined word in the following sentence by selecting the best option:\n\"She loved to describe the escapades of her three brothers...\"",
     "grave acts","funny acts","serious older acts","mischievous acts",
     'D','','Comprehension,CUET English 2025,Miss Stubbs','Escapades = wild, mischievous adventures.'],

    # ── PASSAGE 2: Dance (Q540-543) ───────────────────────────────────────
    ['English','Reading Comprehension','Passage Comprehension','Medium','Medium','CUET','5','-1','MCQ',
     "Passage: If my younger self could see me now, she would be incredulous. That I work in the field of dance or decipher and translate dance for my own comprehension, call it choreography if you wish, would have been unbelievable. In this respect, I am particularly envious of dancers who claim that they are 'born to dance', implying that it was clearly laid out for them from the beginning. I must say, I find this assertion dubious: it is rarely that easy. To dance means to struggle. In truth, as a child, I never did want to dance. It was forced upon me by a doting mother and a silent father. My father probably had his passive disapproval to avoid argument. From the beginning, my lessons took place under trying conditions. I believe that the conditions were more trying for my mother than for me. She travelled in local, over-crowded trains to the dance class with an unwilling child, tired from a whole day at school. Interestingly, when I was seven, we went to see a movie starring Mumbai Ali, who did a dance number. When we arrived home, I began prancing around the house imitating the film actor and my mother, who was quietly watching, was the one who said 'Kunjudi, you are born to dance'. Ironically, I have no recollection of this story; it was my mother who saw this innate ability in me.\n\nQ: The narrator's younger self would be incredulous if she saw the narrator now, as:",
     "she had become a choreographer",
     "she couldn't decipher the nuances of dance",
     "she was too curious to learn dance since her early childhood",
     "her reality and her dreams were not bridged",
     'A','','Comprehension,CUET English 2025,Dance','The narrator works in dance/choreography now — something her younger self would find unbelievable.'],

    ['English','Reading Comprehension','Passage Comprehension','Medium','Medium','CUET','5','-1','MCQ',
     "The narrator says to dance means to 'struggle' as she ___",
     "was envious of the other dancers' talent",
     "had tough time convincing her parents about her passion",
     "doesn't have good memories of her childhood",
     "had to work hard to perfect the art, though unwillingly",
     'D','','Comprehension,CUET English 2025,Dance','She was forced into dance, had to travel under trying conditions and worked hard despite not wanting to dance.'],

    ['English','Reading Comprehension','Passage Comprehension','Medium','Medium','CUET','5','-1','MCQ',
     "What role did her father play in her becoming a dancer?",
     "He gave his passive disapproval",
     "He was indifferent",
     "He was actively involved",
     "He was creating a harmonious atmosphere",
     'A','','Comprehension,CUET English 2025,Dance','The passage states her father "probably had his passive disapproval to avoid argument."'],

    ['English','Reading Comprehension','Passage Comprehension','Medium','Medium','CUET','5','-1','MCQ',
     "The seeds of the narrator's future vocation were sown when she ___",
     "went willingly to the dance class",
     "pranced around the house imitating the dance steps from a movie",
     "was born with an innate ability to dance",
     "was encouraged by her father to take up dancing",
     'B','','Comprehension,CUET English 2025,Dance','After watching Mumbai Ali in a movie, she pranced around imitating the dance steps — her mother then recognised her innate ability.'],

    # ── PASSAGE 3: Food Security (Q544-547) ──────────────────────────────
    ['English','Reading Comprehension','Passage Comprehension','Medium','Medium','CUET','5','-1','MCQ',
     "Passage: The State of Food Security and Nutrition report (2024) revealed that 733 million people faced malnutrition in 2023. Hidden hunger is a deficiency of essential micronutrients like zinc, iodine, and iron. The signs of this form of malnutrition are 'hidden' as individuals may appear healthy while suffering severe health impacts. Clinical signs emerge only in extreme cases. Traditionally linked to caloric deficiency, hunger is now recognised to include micronutrient inadequacy, which can harm health even without overt signs of disease. The Food Security and Nutrition report highlights that while it may seem intuitive that food-insecure individuals are less likely to maintain a healthy diet, the relationship is complex, shaped by factors like food environments, consumer behaviour, and the affordability of nutritious foods. In some cases, food insecurity is linked to lower consumption of all food types and a higher reliance on staple foods for dietary energy. In others, it can be associated with reduced intake of nutritious foods and increased consumption of energy-dense foods high in unhealthy fats, sugars, and salt. As a result, food insecurity and 'hidden hunger' can result not only in undernutrition but can also lead to overweight and obesity.\n\nQ: What is the primary distinction between 'hidden hunger' and traditional caloric deficiency?",
     "Hidden hunger refers to a lack of protein, whereas caloric deficiency refers to a lack of fats",
     "Hidden hunger is caused by excessive food intake, whereas caloric deficiency is caused by food scarcity",
     "Hidden hunger involves not-so-obvious micronutrient deficiencies, while caloric deficiency has visible symptoms",
     "Hidden hunger is a short-term condition, whereas caloric deficiency affects only long-term health",
     'C','','Comprehension,CUET English 2025,Food Security','The passage states hidden hunger is hidden as individuals may appear healthy — clinical signs emerge only in extreme cases, unlike visible caloric deficiency.'],

    ['English','Reading Comprehension','Passage Comprehension','Medium','Medium','CUET','5','-1','MCQ',
     "Which factor complicates the relationship between food insecurity and diet quality?",
     "The absence of global food production data",
     "The overreliance on scientific studies rather than real-world evidence",
     "Variations in food environments, consumer behaviour, and affordability",
     "The assumption that food insecurity always results in undernutrition",
     'C','','Comprehension,CUET English 2025,Food Security','The passage explicitly mentions food environments, consumer behaviour, and affordability as complicating factors.'],

    ['English','Reading Comprehension','Passage Comprehension','Medium','Medium','CUET','5','-1','MCQ',
     "How can food insecurity paradoxically contribute to obesity?",
     "by causing people to eat excessive amounts of protein",
     "by encouraging over-consumption of cheap, unhealthy energy-dense foods",
     "by limiting access to all food types, reducing overall consumption",
     "by increasing the availability of nutrient-rich food in low-income areas",
     'B','','Comprehension,CUET English 2025,Food Security','Food insecurity leads to increased consumption of energy-dense foods high in unhealthy fats, sugars, and salt — which can cause overweight and obesity.'],

    ['English','Reading Comprehension','Passage Comprehension','Medium','Medium','CUET','5','-1','MCQ',
     "What consequence does the World Bank estimate for a 1% rise in global food prices?",
     "A decline in global malnutrition rates",
     "A reduction in food insecurity through economic growth",
     "An increase of 10 million people towards extreme poverty",
     "No significant impact on food security",
     'C','','Comprehension,CUET English 2025,Food Security','The passage states a 1% rise in global food prices could push 10 million people into extreme poverty.'],

    # ── SENTENCE REARRANGEMENT (Q548-554) ────────────────────────────────
    ['English','Sentence Rearrangement','Sentence Formation','Medium','Medium','CUET','5','-1','MCQ',
     "Rearrange the following phrases into a meaningful sentence. The beginning of the sentence has been given below. Read the options and find out which of the four combinations is correct:\nThe need for a scientific understanding...\n(A) the Rio Earth Summit, the first of its kind\n(B) involving world leaders on environment and development\n(C) including forests was the main objective of\n(D) of phenomena like climate change, biodiversity, and conservation",
     "(B), (D), (C), (A)","(C), (D), (B), (A)","(A), (D), (B), (C)","(B), (C), (D), (A)",
     'D','','Sentence Rearrangement,CUET English 2025','Correct order: The need for a scientific understanding (D) of phenomena like climate change... (B) involving world leaders... (C) including forests was the main objective of (A) the Rio Earth Summit.'],

    ['English','Sentence Rearrangement','Sentence Formation','Medium','Medium','CUET','5','-1','MCQ',
     "Rearrange the following phrases into a meaningful sentence. The beginning of the sentence has been given below:\nThe point is that both ecology and resource management schemes...\n(A) and mechanistic worldview, shaped by the utilitarian promise\n(B) of the Industrial age, had more to say\n(C) about the human mission to extract rather than to conserve\n(D) that developed under the conventional",
     "(B), (D), (A), (C)","(D), (A), (B), (C)","(A), (B), (D), (C)","(C), (D), (A), (B)",
     'B','','Sentence Rearrangement,CUET English 2025',''],

    ['English','Sentence Rearrangement','Sentence Formation','Medium','Medium','CUET','5','-1','MCQ',
     "Rearrange the following phrases into a meaningful sentence:\nApart from many external factors like...\n(A) the language communities cannot shirk\n(B) their responsibility for the endangerment of the local languages\n(C) globalisation, and the\n(D) hegemony of English",
     "(C), (D), (A), (B)","(D), (A), (C), (B)","(A), (C), (D), (B)","(C), (A), (D), (B)",
     'A','','Sentence Rearrangement,CUET English 2025',''],

    ['English','Sentence Rearrangement','Sentence Formation','Medium','Medium','CUET','5','-1','MCQ',
     "Rearrange the following phrases into a meaningful sentence:\n(A) In the region\n(B) had percolated into the underground water\n(C) thus posing a water crisis\n(D) the pollutants from the factory",
     "(A), (D), (B), (C)","(D), (B), (A), (C)","(A), (B), (D), (C)","(D), (A), (C), (B)",
     'A','','Sentence Rearrangement,CUET English 2025',''],

    ['English','Sentence Rearrangement','Sentence Formation','Medium','Medium','CUET','5','-1','MCQ',
     "Rearrange the following phrases into a meaningful sentence:\nOne common mistake that ___\n(A) thinking that it\n(B) would be a waste of time\n(C) to work together\n(D) many people have made it",
     "(D), (A), (B), (C)","(C), (D), (A), (B)","(A), (C), (D), (B)","(D), (C), (A), (B)",
     'A','','Sentence Rearrangement,CUET English 2025',''],

    ['English','Sentence Rearrangement','Sentence Formation','Medium','Medium','CUET','5','-1','MCQ',
     "Rearrange the following phrases into a meaningful sentence:\nStimulants such as caffeine ___\n(A) activate adrenal glands\n(B) deplete valuable minerals\n(C) worsening fatigue and sleep issues\n(D) needed for energy and",
     "(A), (D), (B), (C)","(B), (A), (D), (C)","(D), (A), (B), (C)","(A), (B), (C), (D)",
     'A','','Sentence Rearrangement,CUET English 2025',''],

    ['English','Sentence Rearrangement','Sentence Formation','Medium','Medium','CUET','5','-1','MCQ',
     "Rearrange the following phrases into a meaningful sentence:\nPremature greying of the workforce ___\n(A) on productivity and\n(B) long-term economic growth\n(C) has a negative effect\n(D) required for large economies",
     "(C), (A), (B), (D)","(B), (C), (A), (D)","(A), (C), (D), (B)","(C), (D), (A), (B)",
     'A','','Sentence Rearrangement,CUET English 2025',''],

    # ── MATCH LIST — Idioms & Meanings (Q555) ────────────────────────────
    ['English','Idioms and Phrases','Match the Column','Medium','Medium','CUET','5','-1','MCQ',
     "Match the idioms in List-I with their meanings in List-II:\nList-I:\n(P) Put one's shoulder to the wheel\n(Q) Throw down the gauntlet\n(R) Get too big for one's boots\n(S) Cut one's coat according to one's cloth\n\nList-II:\n(I) Work hard at a task\n(II) Take up a challenge\n(III) Become very conceited\n(IV) Spend only what one can afford\n\nChoose the correct answer from the options given below:",
     "(P)-(II), (Q)-(I), (R)-(IV), (S)-(III)","(P)-(I), (Q)-(II), (R)-(III), (S)-(IV)",
     "(P)-(I), (Q)-(IV), (R)-(II), (S)-(III)","(P)-(III), (Q)-(II), (R)-(I), (S)-(IV)",
     'B','','Idioms,Match List,CUET English 2025','P=Work hard; Q=Take up a challenge; R=Become very conceited; S=Spend only what one can afford'],

    # ── MATCH LIST — Word Definitions (Q556) ─────────────────────────────
    ['English','Vocabulary','Match the Column','Medium','Medium','CUET','5','-1','MCQ',
     "Match the words in List-I with their definitions in List-II:\nList-I:\n(P) Expurgate\n(Q) Excerpt\n(R) Plagiarism\n(S) Besiege\n\nList-II:\n(I) To remove offensive portion of a book\n(II) Written draft\n(III) To surround a place with the intention of capturing\n(IV) To give up a throne or other office of dignity\n\nChoose the correct answer from the options given below:",
     "(P)-(I), (Q)-(II), (R)-(III), (S)-(IV)","(P)-(II), (Q)-(I), (R)-(IV), (S)-(III)",
     "(P)-(I), (Q)-(IV), (R)-(II), (S)-(III)","(P)-(IV), (Q)-(I), (R)-(III), (S)-(II)",
     'C','','Vocabulary,Match List,CUET English 2025','Expurgate=remove offensive content(I); Excerpt=extract/draft(II→IV?); Besiege=surround(III)'],

    # ── MATCH LIST — Phrasal Verbs (Q557) ────────────────────────────────
    ['English','Phrasal Verbs','Match the Column','Medium','Medium','CUET','5','-1','MCQ',
     "Match the phrasal verbs in List-I with their meanings in List-II:\nList-I:\n(P) Try one's luck\n(Q) Scale back\n(R) Allude to\n(S) Call upon\n\nList-II:\n(I) To attempt something in an indirect way\n(II) To make production or spending smaller in size or extent\n(III) To hint at something\n(IV) To make something more interesting or exciting\n\nChoose the correct answer:",
     "(P)-(IV), (Q)-(I), (R)-(III), (S)-(II)","(P)-(I), (Q)-(II), (R)-(III), (S)-(IV)",
     "(P)-(III), (Q)-(IV), (R)-(I), (S)-(II)","(P)-(II), (Q)-(I), (R)-(IV), (S)-(III)",
     'B','','Phrasal Verbs,Match List,CUET English 2025','Try one\'s luck=attempt; Scale back=reduce; Allude to=hint at; Call upon=summon/make exciting'],

    # ── FILL IN THE BLANK — Sentences + List (Q558) ──────────────────────
    ['English','Fill in the Blanks','Phrasal Verbs in Context','Medium','Medium','CUET','5','-1','MCQ',
     "Fill in the blanks in the sentences given in List-I with the appropriate phrasal verbs given in List-II:\nList-I (Sentences):\n(A) They _____ that I go with them.\n(B) The oil factory was eventually _____, a vibrant art gallery, preserving its historical significance.\n(C) Before leaving the room, she remembered to _____.\n(D) She ran the air conditioner to keep it cool.\n\nList-II (Phrasal Verbs):\n(I) recommended  (II) turned into  (III) turned up  (IV) turn on\n\nChoose the correct answer:",
     "(A)-(I), (B)-(II), (C)-(III), (D)-(IV)","(A)-(II), (B)-(I), (C)-(IV), (D)-(III)",
     "(A)-(I), (B)-(IV), (C)-(II), (D)-(III)","(A)-(III), (B)-(II), (C)-(IV), (D)-(I)",
     'A','','Fill in the Blanks,Phrasal Verbs,CUET English 2025',''],

    # ── FILL IN THE BLANK — Idioms in sentences (Q559) ───────────────────
    ['English','Fill in the Blanks','Idioms in Context','Medium','Medium','CUET','5','-1','MCQ',
     "Fill in the blanks in the given sentences with the appropriate idioms given in List-II:\nList-I (Sentences):\n(A) With the project deadline approaching, the team had _____ the submission date.\n(B) As the fitness trend started, many people decided to _____ and join the new workout class.\n(C) Despite the promise of secrecy, Frank couldn't resist the temptation to _____ about the surprise party.\n(D) After a long day at work, I am ready to _____ and get a good night's sleep.\n\nList-II (Idioms):\n(I) extended  (II) jump on the bandwagon  (III) let the cat out of the bag  (IV) hit the sack\n\nChoose the correct answer:",
     "(A)-(II), (B)-(I), (C)-(IV), (D)-(III)","(A)-(I), (B)-(II), (C)-(III), (D)-(IV)",
     "(A)-(III), (B)-(IV), (C)-(I), (D)-(II)","(A)-(IV), (B)-(III), (C)-(II), (D)-(I)",
     'B','','Fill in the Blanks,Idioms,CUET English 2025',''],

    # ── FILL IN THE BLANK — Words for culinary personalities (Q560) ──────
    ['English','Fill in the Blanks','Words in Context','Medium','Medium','CUET','5','-1','MCQ',
     "Fill in the blanks in the sentences given in List-I with the appropriate words given in List-II:\nList-I (Sentences):\n(A) Ika Jena is an Odia food _____, who documents and showcases Odia cuisine on her blog Culinary Xpress.\n(B) Gitika Saikia identifies herself as an Assamese food _____ who specialises in tribal and mainland Assamese cuisines.\n(C) Himachal Pradesh has a rich _____ of traditional breads such as siddu (stuffed bread), bhatura (local sourdough flatbread) and pakkan (deep-fried festive flatbread).\n(D) Deepa Chauhan is a Bengaluru-based culinary _____ and Sindhi cuisine specialist.\n\nList-II (Words):\n(I) chef  (II) entrepreneur  (III) chronicles  (IV) repertoire\n\nChoose the correct answer:",
     "(A)-(I), (B)-(II), (C)-(III), (D)-(IV)","(A)-(II), (B)-(I), (C)-(IV), (D)-(III)",
     "(A)-(I), (B)-(IV), (C)-(II), (D)-(III)","(A)-(III), (B)-(II), (C)-(I), (D)-(IV)",
     'A','','Fill in the Blanks,CUET English 2025','Ika Jena=chef; Gitika=entrepreneur; Himachal=chronicles; Deepa=repertoire'],

    # ── FILL IN THE BLANK — Conjunctions (Q561) ──────────────────────────
    ['English','Fill in the Blanks','Conjunctions','Easy','Low','CUET','5','-1','MCQ',
     "Fill in the blanks in the sentences given in List-I with the appropriate words given in List-II:\nList-I (Sentences):\n(A) I read the mystery novels _____ they captivate my imagination.\n(B) This is the place _____ he was born.\n(C) Many things have happened _____ we met last time.\n(D) _____, you tell me the truth, I will not allow you to go.\n\nList-II (Words):\n(I) Unless  (II) because  (III) Where  (IV) Since\n\nChoose the correct answer:",
     "(A)-(II), (B)-(III), (C)-(IV), (D)-(I)","(A)-(I), (B)-(III), (C)-(IV), (D)-(II)",
     "(A)-(IV), (B)-(II), (C)-(I), (D)-(III)","(A)-(III), (B)-(I), (C)-(II), (D)-(IV)",
     'A','','Fill in the Blanks,Conjunctions,CUET English 2025','(A) because; (B) Where; (C) Since; (D) Unless'],

    # ── MATCH — Word Meanings (Q562) ─────────────────────────────────────
    ['English','Vocabulary','Match the Column','Medium','Medium','CUET','5','-1','MCQ',
     "Match the words in List-I with their meanings in List-II:\nList-I:\n(P) Deepen\n(Q) Deplete\n(R) Depreciate\n(S) Deviate\n\nList-II:\n(I) Express disapproval of\n(II) Reduce\n(III) Extreme wickedness\n(IV) Inhabitant\n\nChoose the correct answer from the options given below:",
     "(P)-(III), (Q)-(II), (R)-(I), (S)-(IV)","(P)-(II), (Q)-(III), (R)-(IV), (S)-(I)",
     "(P)-(IV), (Q)-(I), (R)-(II), (S)-(III)","(P)-(I), (Q)-(IV), (R)-(III), (S)-(II)",
     'A','','Vocabulary,Match List,CUET English 2025',''],

    # ── FILL IN THE BLANK — Phrasal verb sentences (Q563) ────────────────
    ['English','Fill in the Blanks','Phrasal Verbs in Context','Medium','Medium','CUET','5','-1','MCQ',
     "Fill in the blanks in the sentences given in List-I with the appropriate phrasal verbs from List-II:\nList-I (Sentences):\n(A) They _____ that I go with them for two days.\n(B) The shoemaker _____ that plans that the rent be changed.\n(C) The landlord _____ that I pay the rent by Tuesday latest.\n(D) _____ (proposed) — he had to decide before the meeting.\n\nList-II (Verbs):\n(I) demanded  (II) extended  (III) insisted  (IV) proposed",
     "(A)-(II), (B)-(IV), (C)-(III), (D)-(I)","(A)-(III), (B)-(I), (C)-(IV), (D)-(II)",
     "(A)-(I), (B)-(II), (C)-(III), (D)-(IV)","(A)-(IV), (B)-(III), (C)-(I), (D)-(II)",
     'A','','Fill in the Blanks,Phrasal Verbs,CUET English 2025',''],

    # ── MATCH — Idioms & Meanings (Q564) ─────────────────────────────────
    ['English','Idioms and Phrases','Match the Column','Medium','Medium','CUET','5','-1','MCQ',
     "Match the idioms in List-I with their meanings in List-II:\nList-I:\n(P) Break the ice\n(Q) Hit the nail on the head\n(R) Plagiarism / Beat around the bush\n(S) Bite the bullet\n\nList-II:\n(I) to describe precisely the main point or issue, addressing it directly\n(II) to deal with a painful/difficult situation with courage and resilience\n(III) to initiate conversation in a social setting especially in a formal or awkward situation\n(IV) something very expensive or costly\n\nChoose the correct answer:",
     "(P)-(III), (Q)-(I), (R)-(IV), (S)-(II)","(P)-(I), (Q)-(III), (R)-(II), (S)-(IV)",
     "(P)-(IV), (Q)-(II), (R)-(I), (S)-(III)","(P)-(II), (Q)-(IV), (R)-(III), (S)-(I)",
     'A','','Idioms,Match List,CUET English 2025','Break the ice=initiate conversation; Hit nail on head=precise; Bite bullet=deal with difficulty'],

    # ── FILL IN THE BLANK — Word Form (Q565) ─────────────────────────────
    ['English','Grammar','Word Forms','Easy','Low','CUET','5','-1','MCQ',
     "Use the correct form of the word 'RESOLVE' to fill in the blank:\nDid you make any New Year's _____ this year?",
     "resolve","resolving","resolution","resolvation",
     'C','','Grammar,Word Forms,CUET English 2025','The noun form "resolution" is correct in this context.'],

    # ── VOCABULARY IN CONTEXT (Q566) ─────────────────────────────────────
    ['English','Vocabulary','Word Meaning in Context','Medium','Medium','CUET','5','-1','MCQ',
     "Choose the appropriate word for the meaning of the underlined word:\nRavi came up on the stage and started speaking impromptu.",
     "Quickly","Unrehearsed","Incoherently","Impressively",
     'B','','Vocabulary,CUET English 2025','Impromptu = done without preparation; unrehearsed.'],

    # ── FILL IN THE BLANK — Adverb (Q567) ────────────────────────────────
    ['English','Fill in the Blanks','Adverbs','Easy','Low','CUET','5','-1','MCQ',
     "Fill in the blank with the correct word from the options given below to make a meaningful sentence:\nTen tourists were _____ injured in the accident and a few of them succumbed to their injuries on the way to the hospital.",
     "spuriously","gravely","fatally","vitally",
     'B','','Fill in the Blanks,CUET English 2025','Gravely = seriously/severely. The sentence implies serious but not necessarily fatal injuries.'],

    # ── FILL IN THE BLANK — Verb (Q568) ──────────────────────────────────
    ['English','Fill in the Blanks','Vocabulary in Context','Easy','Low','CUET','5','-1','MCQ',
     "Fill in the blank with the correct word from the options given below to make a meaningful sentence:\nIn my haste to stop the boiling milk from overflowing, I accidentally _____ my fingers.",
     "scalded","scorched","scathed","scalded (burned by steam/liquid)",
     'A','','Fill in the Blanks,CUET English 2025','Scalded = burned by boiling liquid or steam.'],

    # ── FILL IN THE BLANK — Idiom in context (Q569) ──────────────────────
    ['English','Fill in the Blanks','Prepositions','Easy','Low','CUET','5','-1','MCQ',
     "Fill in the blank with the correct option to make a meaningful sentence:\nHis salary is in _____ with the amount of work he does.",
     "accord","coordination","balance","keeping",
     'D','','Fill in the Blanks,CUET English 2025','"In keeping with" = consistent with / in proportion to.'],

    # ── FILL IN THE BLANK — Adjective (Q570) ─────────────────────────────
    ['English','Fill in the Blanks','Vocabulary in Context','Medium','Medium','CUET','5','-1','MCQ',
     "Fill in the blank with the correct word from the options given below:\nSara was so _____ that she did not take much time to make up her mind.",
     "imperious","impetuous","inarticulate","impertinent",
     'B','','Fill in the Blanks,CUET English 2025','Impetuous = acting or done quickly and without thought; quick to decide.'],

    # ── FILL IN THE BLANK — Preposition (Q571) ───────────────────────────
    ['English','Fill in the Blanks','Prepositions','Easy','Low','CUET','5','-1','MCQ',
     "Fill in the blank with the correct option to make a meaningful sentence:\nShe is standing _____ her husband in this crisis.",
     "with","of","for","by",
     'D','','Fill in the Blanks,CUET English 2025','"Standing by" = supporting someone through difficult times.'],

    # ── FILL IN THE BLANK — Phrasal verb (Q572) ──────────────────────────
    ['English','Fill in the Blanks','Phrasal Verbs','Easy','Low','CUET','5','-1','MCQ',
     "Fill in the blank with the correct option:\nThe two friends appear to have fallen _____ over a minor issue.",
     "down","out","into","by",
     'B','','Fill in the Blanks,Phrasal Verbs,CUET English 2025','"Fall out" = to quarrel or have a disagreement.'],

    # ── FILL IN THE BLANK — Verb (Q573) ──────────────────────────────────
    ['English','Fill in the Blanks','Vocabulary','Medium','Medium','CUET','5','-1','MCQ',
     "Fill in the blank with the correct option to make a meaningful sentence:\nThe committee's appeal to the people for money _____ little response.",
     "evoked","provided","provoked","presented",
     'A','','Fill in the Blanks,CUET English 2025','"Evoked" = drew out or produced (a response).'],

    # ── VOCABULARY — Word Meaning (Q574) ─────────────────────────────────
    ['English','Vocabulary','Word Meaning','Medium','Medium','CUET','5','-1','MCQ',
     "Which of the following best explains the meaning of the word METAMORPHOSIS?",
     "A man flaunts his new car to impress his neighbours",
     "A woman throws out old clothes from her wardrobe",
     "A garden full of weeds is converted into a lovely green lawn",
     "A man paints his new home in bright colours",
     'C','','Vocabulary,CUET English 2025','Metamorphosis = a complete transformation in form, nature or character.'],

    # ── ANTONYMS (Q575-Q578) ──────────────────────────────────────────────
    ['English','Vocabulary','Antonyms','Medium','Medium','CUET','5','-1','MCQ',
     "The patient will certainly recuperate under medical care.\nThe ANTONYM of the word RECUPERATE is:",
     "Resuscitate","Reimburse","Degenerate","Convalesce",
     'C','','Vocabulary,Antonyms,CUET English 2025','Recuperate = recover from illness. Antonym = Degenerate (to deteriorate/worsen).'],

    ['English','Vocabulary','Antonyms','Medium','Medium','CUET','5','-1','MCQ',
     "The speech given by him seemed to be INNOCUOUS but there were undercurrents that implied the contrary.\nThe ANTONYM of the word INNOCUOUS is:",
     "Ineffective","Harmful","Gentle","Insightful",
     'B','','Vocabulary,Antonyms,CUET English 2025','Innocuous = harmless. Antonym = Harmful.'],

    ['English','Vocabulary','Antonyms','Medium','Medium','CUET','5','-1','MCQ',
     "Select the word that is the exact opposite of MELLIFLUOUS in meaning from the given options:",
     "Resonant","Cacophonous","Harmonious","Acrimonious",
     'B','','Vocabulary,Antonyms,CUET English 2025','Mellifluous = sweet/smooth-sounding. Antonym = Cacophonous (harsh/discordant sound).'],

    # ── SYNONYMS (Q579-Q581) ──────────────────────────────────────────────
    ['English','Vocabulary','Synonyms','Medium','Medium','CUET','5','-1','MCQ',
     "Select the word from the given options that is most similar in meaning to the word JUXTAPOSE:",
     "Collective","Adjacent","Aligned","Closest",
     'B','','Vocabulary,Synonyms,CUET English 2025','Juxtapose = place side by side for comparison. Most similar = Adjacent (side by side).'],

    ['English','Vocabulary','Synonyms','Medium','Medium','CUET','5','-1','MCQ',
     "Select the word from the given options that is most similar in meaning to the word EMULATE:",
     "Echo","Rehearse","Envy","Mimic",
     'D','','Vocabulary,Synonyms,CUET English 2025','Emulate = match or surpass by imitation. Synonym = Mimic.'],

    ['English','Vocabulary','Synonyms','Medium','Medium','CUET','5','-1','MCQ',
     "Select the word from the given options that is most similar in meaning to the word ALLEVIATE:",
     "Swell","Aggravate","Royal","Assuage",
     'D','','Vocabulary,Synonyms,CUET English 2025','Alleviate = make less severe. Synonym = Assuage.'],

    # ── ANTONYMS (Q582-Q585) ──────────────────────────────────────────────
    ['English','Vocabulary','Antonyms','Medium','Medium','CUET','5','-1','MCQ',
     "Choose the word from the given options which is opposite in meaning to the given word: INSIPID",
     "Uninviting","Savory","Spiritless","Spirited",
     'B','','Vocabulary,Antonyms,CUET English 2025','Insipid = lacking flavour or dull. Antonym = Savory (strong, pleasant taste).'],

    ['English','Vocabulary','Antonyms','Medium','Medium','CUET','5','-1','MCQ',
     "Choose the word from the given options which is opposite in meaning to the given word: PRAGMATIC",
     "Practical","Sensible","Proficient","Utopian",
     'D','','Vocabulary,Antonyms,CUET English 2025','Pragmatic = dealing with things practically. Antonym = Utopian (idealistic, unrealistic).'],

    ['English','Vocabulary','Synonyms','Medium','Medium','CUET','5','-1','MCQ',
     "Choose the word from the given options that is similar in meaning to the given word: ANTICIPATE",
     "Contemplate","Assert","Neglect","Predict",
     'D','','Vocabulary,Synonyms,CUET English 2025','Anticipate = expect or look forward to. Synonym = Predict.'],

    ['English','Vocabulary','Antonyms','Medium','Medium','CUET','5','-1','MCQ',
     "Choose the correct antonym for the word 'Flagrant' from the options given below:",
     "Notorious","Blatant","Unobtrusive","Adjacent",
     'C','','Vocabulary,Antonyms,CUET English 2025','Flagrant = conspicuously offensive or wrong. Antonym = Unobtrusive (not conspicuous/noticeable).'],
]

for row_data in questions:
    ws.append(row_data)

col_widths = [15, 25, 25, 12, 12, 8, 6, 12, 10, 80, 45, 45, 45, 45, 14, 14, 35, 50]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[ws.cell(1, i).column_letter].width = width

for row_idx in range(2, len(questions) + 2):
    fill_color = "EBF3FB" if row_idx % 2 == 0 else "FFFFFF"
    fill = PatternFill("solid", fgColor=fill_color)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill = fill

ws.freeze_panes = "A2"

output_path = "D:/Bkp/GRID/GridAcademy/sample/CUET_English_2025_Questions.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Total questions written: {len(questions)}")
